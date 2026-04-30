import os
import tempfile
import time
from pathlib import Path

from fpdf import FPDF

from core.export_utils import pdf_output_bytes, safe_text
from core.utils import decodificar_base64_seguro, mapa_detalles_pacientes
from core.app_logging import log_event

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"

from core._exports_helpers import (
    collect_patient_sections as _collect_patient_sections_impl,
    collect_sql_sections as _collect_sql_sections,
    local_section_records as _local_section_records,
    merge_records as _merge_records,
    patient_context as _patient_context,
    sql_sections_empty as _sql_sections_empty,
    patient_signature_bytes as _patient_signature_bytes,
    doctor_signature_bytes as _doctor_signature_bytes,
    order_attachment_note as _order_attachment_note,
)
from core._exports_pdf_base import (
    pdf_header_oscuro as _pdf_header_oscuro,
    section_title as _section_title,
    write_multiline_text as _write_multiline_text,
    write_pairs as _write_pairs,
    RespaldoClinicoPDF,
)
from core._exports_excel_json import (
    build_patient_excel_bytes as _build_patient_excel_bytes_impl,
    build_patient_json_bytes as _build_patient_json_bytes_impl,
)
from core._exports_history import (
    REPORTLAB_DISPONIBLE,
    build_history_pdf_bytes as _build_history_pdf_bytes,
)
from core._exports_backup import build_backup_pdf_bytes


def collect_patient_sections(session_state, paciente_sel):
    """
    Wrapper compatible con tests viejos que monkeypatchean helpers en `core.clinical_exports`.
    """
    _ck_secs = f"_ce_secs_{paciente_sel}"
    _secs_cached = session_state.get(_ck_secs)
    local_ts = session_state.get("_ultimo_guardado_ts", 0)
    if (
        _secs_cached is not None
        and isinstance(_secs_cached, dict)
        and _secs_cached.get("local_ts") == local_ts
        and time.monotonic() - _secs_cached.get("ts", 0) < 30
    ):
        return _secs_cached["sections"]

    ctx = _patient_context(session_state, paciente_sel)
    local_sections = {
        "Auditoria de Presencia": _local_section_records(session_state, "checkin_db", paciente_sel, ctx),
        "Visitas y Agenda": _local_section_records(session_state, "agenda_db", paciente_sel, ctx),
        "Emergencias y Ambulancia": _local_section_records(session_state, "emergencias_db", paciente_sel, ctx),
        "Enfermeria y Plan de Cuidados": _local_section_records(session_state, "cuidados_enfermeria_db", paciente_sel, ctx),
        "Escalas Clinicas": _local_section_records(session_state, "escalas_clinicas_db", paciente_sel, ctx),
        "Auditoria Legal": _local_section_records(session_state, "auditoria_legal_db", paciente_sel, ctx),
        "Procedimientos y Evoluciones": _local_section_records(session_state, "evoluciones_db", paciente_sel, ctx),
        "Estudios Complementarios": _local_section_records(session_state, "estudios_db", paciente_sel, ctx),
        "Materiales Utilizados": _local_section_records(session_state, "consumos_db", paciente_sel, ctx),
        "Registro de Heridas": _local_section_records(session_state, "fotos_heridas_db", paciente_sel, ctx),
        "Signos Vitales": _local_section_records(session_state, "vitales_db", paciente_sel, ctx),
        "Control Pediatrico": _local_section_records(session_state, "pediatria_db", paciente_sel, ctx),
        "Balance Hidrico": _local_section_records(session_state, "balance_db", paciente_sel, ctx),
        "Plan Terapeutico": _local_section_records(session_state, "indicaciones_db", paciente_sel, ctx),
        "Consentimientos": _local_section_records(session_state, "consentimientos_db", paciente_sel, ctx),
        "Cobros y Facturacion": _local_section_records(session_state, "facturacion_db", paciente_sel, ctx),
    }
    sql_sections = _collect_sql_sections(session_state, paciente_sel, ctx)
    merged = {
        nombre: _merge_records(local_sections.get(nombre, []), sql_sections.get(nombre, []))
        for nombre in (
            "Auditoria de Presencia",
            "Visitas y Agenda",
            "Emergencias y Ambulancia",
            "Enfermeria y Plan de Cuidados",
            "Escalas Clinicas",
            "Auditoria Legal",
            "Procedimientos y Evoluciones",
            "Estudios Complementarios",
            "Materiales Utilizados",
            "Registro de Heridas",
            "Signos Vitales",
            "Control Pediatrico",
            "Balance Hidrico",
            "Plan Terapeutico",
            "Consentimientos",
            "Cobros y Facturacion",
        )
    }
    try:
        session_state[_ck_secs] = {"ts": time.monotonic(), "local_ts": local_ts, "sections": merged}
    except Exception as _exc:
        log_event("clinical_exports", f"fallo_cache_secciones:{type(_exc).__name__}")
    return merged


def build_patient_excel_bytes(session_state, paciente_sel):
    from datetime import datetime as _dt
    import io

    import pandas as pd

    detalles = mapa_detalles_pacientes(session_state).get(paciente_sel, {})
    nom_pac = paciente_sel.split(" - ")[0] if " - " in paciente_sel else paciente_sel

    output = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        engine = "openpyxl"
    except Exception:
        try:
            import xlsxwriter  # noqa: F401

            engine = "xlsxwriter"
        except Exception:
            return None

    sheets = {}
    for section_name, records in collect_patient_sections(session_state, paciente_sel).items():
        if not records:
            continue
        df = pd.DataFrame(records).drop(columns=["paciente", "empresa", "firma_b64"], errors="ignore")
        sheets[section_name[:31]] = df

    if not sheets:
        sheets["Sin registros"] = pd.DataFrame([{"mensaje": "No hay registros clinicos para este paciente."}])

    with pd.ExcelWriter(output, engine=engine) as writer:
        resumen_rows = [{"Seccion": sec, "Registros": len(df)} for sec, df in sheets.items()]
        pd.DataFrame(resumen_rows).to_excel(writer, index=False, sheet_name="Resumen")
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        if engine == "openpyxl":
            wb = writer.book
            _navy = "162644"
            _green = "0D5A50"
            _hdr_font = Font(bold=True, color="FFFFFF", size=10)
            _hdr_fill_navy = PatternFill("solid", fgColor=_navy)
            _hdr_fill_green = PatternFill("solid", fgColor=_green)
            _thin = Side(style="thin", color="D1D5DB")
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

            for ws in wb.worksheets:
                fill = _hdr_fill_green if ws.title == "Resumen" else _hdr_fill_navy
                for cell in ws[1]:
                    cell.font = _hdr_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = _border
                ws.row_dimensions[1].height = 22
                for col in ws.columns:
                    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 50)
                alt_fill = PatternFill("solid", fgColor="F3F4F6")
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    row_fill = alt_fill if i % 2 == 0 else None
                    for cell in row:
                        cell.border = _border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        if row_fill:
                            cell.fill = row_fill
                ws.freeze_panes = "A2"

            ws_res = wb["Resumen"]
            ws_res.insert_rows(1, amount=4)
            ws_res["A1"] = detalles.get("empresa", "")
            ws_res["A1"].font = Font(bold=True, size=14, color=_navy)
            ws_res["A2"] = "HISTORIA CLINICA DIGITAL - EXPORTACION COMPLETA"
            ws_res["A2"].font = Font(bold=True, size=11, color=_navy)
            ws_res["A3"] = (
                f"Paciente: {nom_pac}   |   DNI: {detalles.get('dni', 'S/D')}   |   Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
            )
            ws_res["A3"].font = Font(size=9, color="64748B")
            ws_res["A4"] = ""

    output.seek(0)
    return output.getvalue()


def build_patient_json_bytes(session_state, paciente_sel):
    import json

    payload = {
        "paciente": paciente_sel,
        "detalles": mapa_detalles_pacientes(session_state).get(paciente_sel, {}),
        "secciones": collect_patient_sections(session_state, paciente_sel),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def build_history_pdf_bytes(session_state, paciente_sel, mi_empresa, profesional=None):
    if not REPORTLAB_DISPONIBLE:
        return None
    return _build_history_pdf_bytes(
        session_state,
        paciente_sel,
        mi_empresa,
        profesional=profesional,
    )


def build_consent_pdf_bytes(session_state, paciente_sel, mi_empresa, profesional=None):
    consentimientos = [x for x in session_state.get("consentimientos_db", []) if x.get("paciente") == paciente_sel]
    if not consentimientos:
        return None

    consentimiento = consentimientos[-1]
    detalles = mapa_detalles_pacientes(session_state).get(paciente_sel, {})

    pdf = FPDF()
    pdf.set_margins(14, 12, 14)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    _pdf_header_oscuro(
        pdf, detalles.get("empresa", mi_empresa),
        "CONSENTIMIENTO INFORMADO",
        subtitulo="Atencion y Terapia en Domicilio",
        badge_txt="Consentimiento",
        badge_rgb=(13, 90, 80),
    )
    pdf.ln(4)

    texto = (
        "Por medio del presente, dejo constancia de que he recibido informacion clara, suficiente y comprensible "
        "respecto de la modalidad de atencion y/o terapia en domicilio propuesta para el paciente. "
        "Se me explicaron los objetivos asistenciales, la frecuencia estimada de controles, el alcance de las "
        "prestaciones, los cuidados generales esperables, las pautas de alarma clinica y la necesidad de mantener "
        "condiciones adecuadas de acceso, higiene y seguridad para el desarrollo de la atencion. "
        "\n\n"
        "En consecuencia, en mi caracter de paciente y/o familiar responsable, presto conformidad para recibir "
        "atencion sanitaria, controles, curaciones, seguimiento clinico y/o terapia en el domicilio informado, "
        "autorizando asimismo el registro clinico, administrativo y documental correspondiente dentro del sistema."
        "\n\n"
        "Declaro que los datos aportados son veridicos y que he podido realizar preguntas, recibiendo respuesta "
        "satisfactoria por parte del profesional interviniente."
    )
    pdf.set_font("Arial", "", 11)
    _write_multiline_text(pdf, texto, line_height=7)
    pdf.ln(6)

    _write_pairs(
        pdf,
        [
            ("Paciente", paciente_sel),
            ("DNI paciente", detalles.get("dni", "S/D")),
            ("Domicilio", detalles.get("direccion", "S/D")),
            ("Fecha", consentimiento.get("fecha", "S/D")),
            ("Firmante", consentimiento.get("firmante", paciente_sel)),
            ("DNI firmante", consentimiento.get("dni_firmante", detalles.get("dni", "S/D"))),
            ("Vinculo", consentimiento.get("vinculo", "Paciente")),
            ("Telefono", consentimiento.get("telefono", detalles.get("telefono", "S/D"))),
            ("Observaciones", consentimiento.get("observaciones", "")),
            ("Profesional responsable", (profesional or {}).get("nombre", "S/D")),
            ("Matricula", (profesional or {}).get("matricula", "S/D")),
        ],
    )

    pdf.ln(6)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, safe_text("Clausulas de conformidad"), ln=True)
    pdf.set_font("Arial", "", 9)
    clausulas = [
        "1. El firmante acepta la atencion domiciliaria en el domicilio consignado.",
        "2. El firmante se compromete a informar cambios clinicos relevantes y dificultades de acceso.",
        "3. La firma acredita conformidad con la modalidad de prestacion y con el registro documental.",
    ]
    for clausula in clausulas:
        _write_multiline_text(pdf, clausula, line_height=5)

    firma_bytes = None
    if consentimiento.get("firma_b64"):
        firma_bytes = decodificar_base64_seguro(consentimiento["firma_b64"]) or None
    if firma_bytes:
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                tmp.write(firma_bytes)
                tmp_path = tmp.name
            pdf.image(tmp_path, x=30, y=210, w=60)
        except Exception as _exc:
            log_event("clinical_exports", f"fallo_firma_consentimiento:{type(_exc).__name__}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)

    y_firma = max(pdf.get_y() + 24, 240)
    if y_firma > 262:
        pdf.add_page()
        y_firma = 230

    pdf.line(25, y_firma, 90, y_firma)
    pdf.set_xy(25, y_firma + 2)
    pdf.set_font("Arial", "", 9)
    pdf.cell(65, 5, safe_text("Firma paciente / familiar"), align="C")

    pdf.line(120, y_firma, 185, y_firma)
    pdf.set_xy(120, y_firma + 2)
    pdf.cell(65, 5, safe_text("Firma y sello profesional"), align="C")

    return pdf_output_bytes(pdf)


def build_prescription_pdf_bytes(session_state, paciente_sel, mi_empresa, record, profesional=None):
    detalles = mapa_detalles_pacientes(session_state).get(paciente_sel, {})
    estado_actual = record.get("estado_receta") or record.get("estado_clinico") or "Activa"
    medico_indicante = record.get("medico_nombre") or (profesional or {}).get("nombre", "S/D")
    matricula_indicante = record.get("medico_matricula") or (profesional or {}).get("matricula", "S/D")

    _estado_colors = {
        "Activa": (13, 90, 80),
        "Suspendida": (180, 60, 20),
        "Completada": (60, 80, 120),
        "Modificada": (130, 100, 0),
    }
    _badge_color = _estado_colors.get(estado_actual, (60, 80, 120))

    pdf = FPDF()
    pdf.set_margins(14, 12, 14)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    _pdf_header_oscuro(
        pdf, detalles.get("empresa", mi_empresa),
        "PRESCRIPCION Y ACTA LEGAL DE MEDICACION",
        subtitulo=safe_text(f"Paciente: {paciente_sel.split(' - ')[0]}  ·  Medico: {medico_indicante}  ·  Mat: {matricula_indicante}"),
        badge_txt=estado_actual,
        badge_rgb=_badge_color,
    )

    _section_title(pdf, "Datos del paciente")
    _write_pairs(
        pdf,
        [
            ("Paciente", paciente_sel),
            ("DNI", detalles.get("dni", "S/D")),
            ("Fecha de nacimiento", detalles.get("fnac", "S/D")),
            ("Obra social", detalles.get("obra_social", "S/D")),
            ("Domicilio", detalles.get("direccion", "S/D")),
            ("Telefono", detalles.get("telefono", "S/D")),
            ("Alergias", detalles.get("alergias", "Sin datos")),
        ],
    )

    _section_title(pdf, "Indicacion medica")
    _write_pairs(
        pdf,
        [
            ("Fecha de indicacion", record.get("fecha", "S/D")),
            ("Medicacion / Indicacion", record.get("med", "S/D")),
            ("Estado actual", estado_actual),
            ("Origen del registro", record.get("origen_registro", "Prescripcion digital")),
            ("Medico indicante", medico_indicante),
            ("Matricula profesional", matricula_indicante),
            ("Registrado por", record.get("firmado_por", "S/D")),
        ],
    )
    nota_adjunto = _order_attachment_note(record)
    if nota_adjunto:
        _write_pairs(pdf, [("Adjunto legal", nota_adjunto)])

    if estado_actual != "Activa":
        _section_title(pdf, "Acta de cambio terapeutico")
        _write_pairs(
            pdf,
            [
                ("Tipo de accion", estado_actual),
                ("Fecha y hora del cambio", record.get("fecha_estado") or record.get("fecha_suspension", "S/D")),
                ("Profesional interviniente", record.get("profesional_estado", "S/D")),
                ("Matricula del interviniente", record.get("matricula_estado", "S/D")),
                ("Motivo medico / legal", record.get("motivo_estado", "Sin detalle consignado")),
            ],
        )
        pdf.ln(2)
        pdf.set_font("Arial", "B", 10)
        _write_multiline_text(
            pdf,
            (
                "La presente constancia deja asentado el cambio de estado de la indicacion farmacologica dentro del "
                "registro clinico institucional, con fecha, hora y profesional responsable."
            ),
            line_height=6,
        )

    firma_medica = _doctor_signature_bytes(record)
    if firma_medica:
        _section_title(pdf, "Firma del profesional")
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                tmp.write(firma_medica)
                tmp_path = tmp.name
            y_img = pdf.get_y() + 2
            pdf.image(tmp_path, x=20, y=y_img, w=60)
            pdf.set_xy(90, y_img + 8)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 6, safe_text(medico_indicante), ln=True)
            pdf.set_x(90)
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 6, safe_text(f"Matricula: {matricula_indicante}"), ln=True)
            pdf.set_x(90)
            pdf.cell(0, 6, safe_text(f"Fecha de firma: {record.get('fecha', 'S/D')}"), ln=True)
            pdf.ln(34)
        except Exception as _exc:
            log_event("clinical_exports", f"fallo_firma_medica:{type(_exc).__name__}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        _section_title(pdf, "Firma del profesional")
        _write_pairs(
            pdf,
            [
                ("Profesional", medico_indicante),
                ("Matricula", matricula_indicante),
                ("Firma digital", "No disponible en el registro"),
            ],
        )

    pdf.ln(8)
    y_base = max(pdf.get_y() + 12, 240)
    if y_base > 262:
        pdf.add_page()
        y_base = 235

    pdf.line(20, y_base, 90, y_base)
    pdf.set_xy(20, y_base + 2)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(70, 5, safe_text(medico_indicante), ln=True, align="C")
    pdf.set_x(20)
    pdf.set_font("Arial", "", 9)
    pdf.cell(70, 5, safe_text(f"Matricula: {matricula_indicante}"), ln=True, align="C")

    pdf.line(120, y_base, 190, y_base)
    pdf.set_xy(120, y_base + 2)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(70, 5, safe_text("Paciente / Familiar notificado"), ln=True, align="C")
    pdf.set_x(120)
    pdf.set_font("Arial", "", 9)
    pdf.cell(70, 5, safe_text(f"Aclaracion: {paciente_sel.split(' - ')[0]}"), ln=True, align="C")
    pdf.set_x(120)
    pdf.cell(70, 5, safe_text(f"DNI: {detalles.get('dni', 'S/D')}"), ln=True, align="C")

    return pdf_output_bytes(pdf)


def build_emergency_pdf_bytes(session_state, paciente_sel, mi_empresa, record, profesional=None):
    """PDF compacto de emergencia — diseñado para caber en 1 pagina A4."""
    detalles = mapa_detalles_pacientes(session_state).get(paciente_sel, {})
    firma_b64 = record.get("firma_b64", "")
    firma_bytes = decodificar_base64_seguro(firma_b64) or None if firma_b64 else None
    prof_nombre = record.get("profesional") or (profesional or {}).get("nombre", "")
    prof_mat = record.get("matricula") or (profesional or {}).get("matricula", "")

    _SKIP = {"", None, "S/D", "Sin traslado confirmado", "0", 0}
    _triage_colors = {
        "Grado 1 - Rojo":    (192, 38,  38),
        "Grado 2 - Amarillo":(180, 120,  0),
        "Grado 3 - Verde":   ( 22, 120, 60),
    }
    triage_grado = record.get("triage_grado", "")
    triage_rgb = _triage_colors.get(triage_grado, (60, 80, 120))

    def _v(val, fallback=""):
        return str(val).strip() if val not in _SKIP else fallback

    def _row(pdf, label, value, lw=48, lh=5):
        txt = _v(value)
        if not txt:
            return
        pdf.set_font("Arial", "B", 8)
        pdf.set_x(pdf.l_margin)
        pdf.cell(lw, lh, safe_text(str(label) + ":"), border=0)
        pdf.set_font("Arial", "", 8)
        remaining = pdf.w - pdf.l_margin - pdf.r_margin - lw
        pdf.multi_cell(remaining, lh, safe_text(txt), border=0)

    def _inline_pairs(pdf, pairs, cols=2):
        usable = pdf.w - pdf.l_margin - pdf.r_margin
        col_w = usable / cols
        items = [(l, _v(v)) for l, v in pairs if _v(v)]
        for i in range(0, len(items), cols):
            row_items = items[i:i + cols]
            x_start = pdf.l_margin
            for j, (lbl, val) in enumerate(row_items):
                pdf.set_xy(x_start + j * col_w, pdf.get_y())
                pdf.set_font("Arial", "B", 8)
                pdf.set_text_color(80, 80, 100)
                pdf.cell(28, 5, safe_text(str(lbl) + ":"), border=0)
                pdf.set_font("Arial", "", 8)
                pdf.set_text_color(20, 20, 20)
                pdf.cell(col_w - 28, 5, safe_text(val), border=0)
            pdf.ln(5)
        pdf.set_text_color(0, 0, 0)

    def _sec(pdf, title):
        if pdf.get_y() > 262:
            pdf.add_page()
        pdf.ln(2)
        pdf.set_fill_color(35, 55, 90)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 8)
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, 6, safe_text("  " + title.upper()), ln=True, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    pdf = FPDF()
    pdf.set_margins(14, 12, 14)
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    # ── Cabecera: bloque oscuro completo ──────────────────────────────
    header_h = 36
    # Fondo principal azul oscuro
    pdf.set_fill_color(22, 38, 68)
    pdf.rect(0, 0, pdf.w, header_h, "F")
    # Franja de color de triage en el margen izquierdo
    pdf.set_fill_color(*triage_rgb)
    pdf.rect(0, 0, 5, header_h, "F")

    # Logo dentro del header
    logo_y = 5
    logo_w = 26
    for ruta in [
        ASSETS_DIR / "logo_medicare_pro.jpeg",
        ASSETS_DIR / "logo_medicare_pro.jpg",
        ASSETS_DIR / "logo_medicare_pro.png",
    ]:
        if ruta.exists():
            try:
                pdf.image(str(ruta), x=8, y=logo_y, w=logo_w)
            except Exception as _exc:
                log_event("clinical_exports", f"fallo_logo_pdf:{type(_exc).__name__}:{ruta.name}")
            break

    # Empresa + título en blanco
    pdf.set_xy(38, 7)
    pdf.set_font("Arial", "B", 14)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 7, safe_text(detalles.get("empresa", mi_empresa)), ln=True)
    pdf.set_x(38)
    pdf.set_font("Arial", "B", 8)
    pdf.set_text_color(160, 200, 255)
    pdf.cell(0, 5, safe_text("ACTA DE EMERGENCIA Y TRASLADO"), ln=True)
    pdf.set_x(38)
    pdf.set_font("Arial", "", 7)
    pdf.set_text_color(200, 210, 230)
    pdf.cell(0, 5, safe_text(
        f"Fecha: {record.get('fecha_evento','')} {record.get('hora_evento','')}   "
        f"Profesional: {prof_nombre}   Mat: {prof_mat}"
    ), ln=True)

    # Badge de triage (esquina superior derecha dentro del header)
    badge_w = 52
    badge_x = pdf.w - badge_w - 6
    pdf.set_fill_color(*triage_rgb)
    pdf.rect(badge_x, 10, badge_w, 16, "F")
    pdf.set_font("Arial", "B", 10)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(badge_x, 12)
    pdf.cell(badge_w, 6, safe_text(triage_grado or "Sin triage"), align="C", border=0, ln=True)
    pdf.set_font("Arial", "", 7)
    pdf.set_xy(badge_x, 19)
    pdf.cell(badge_w, 5, safe_text(_v(record.get("prioridad", ""))), align="C", border=0)

    # Resetear colores y posicion
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(pdf.l_margin, header_h + 4)

    # --- Paciente (1 fila) ---
    _sec(pdf, "Paciente")
    _inline_pairs(pdf, [
        ("Paciente", paciente_sel.split(" - ")[0] if " - " in paciente_sel else paciente_sel),
        ("DNI", detalles.get("dni", "")),
        ("F.Nac", detalles.get("fnac", "")),
        ("Obra Social", detalles.get("obra_social", "")),
        ("Domicilio", record.get("direccion_evento") or detalles.get("direccion", "")),
        ("Telefono", detalles.get("telefono", "")),
    ], cols=2)

    # --- Evento ---
    _sec(pdf, "Evento critico")
    _inline_pairs(pdf, [
        ("Categoria", record.get("categoria_evento", "")),
        ("Tipo", record.get("tipo_evento", "")),
        ("Profesional", prof_nombre),
        ("Matricula", prof_mat),
    ], cols=2)
    motivo_txt = _v(record.get("motivo", ""))
    if motivo_txt:
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Arial", "B", 8)
        pdf.cell(28, 5, safe_text("Motivo:"), border=0)
        pdf.set_font("Arial", "", 8)
        pdf.multi_cell(pdf.w - pdf.l_margin - pdf.r_margin - 28, 5, safe_text(motivo_txt), border=0)
        pdf.ln(1)

    # --- Signos vitales (1 fila si hay datos) ---
    vitales = [
        ("TA", record.get("presion_arterial", "")),
        ("FC", record.get("fc", "")),
        ("SaO2", record.get("saturacion", "")),
        ("Temp", record.get("temperatura", "")),
        ("Glucemia", record.get("glucemia", "")),
        ("EVA", record.get("dolor", "")),
        ("Conciencia", record.get("conciencia", "")),
    ]
    vitales_con_dato = [(l, v) for l, v in vitales if _v(v)]
    if vitales_con_dato:
        _sec(pdf, "Signos vitales")
        _inline_pairs(pdf, vitales_con_dato, cols=4)

    # --- Ambulancia ---
    amb_data = [
        ("Ambulancia", "Si" if record.get("ambulancia_solicitada") else "No"),
        ("Movil", record.get("movil", "")),
        ("Destino", record.get("destino", "")),
        ("Tipo traslado", record.get("tipo_traslado", "")),
        ("Solicitud", record.get("hora_solicitud", "")),
        ("Arribo", record.get("hora_arribo", "")),
        ("Salida", record.get("hora_salida", "")),
        ("Receptor", record.get("receptor", "")),
        ("Familiar", record.get("familiar_notificado", "")),
    ]
    if any(_v(v) for _, v in amb_data):
        _sec(pdf, "Ambulancia y traslado")
        _inline_pairs(pdf, amb_data, cols=3)

    # --- Parte asistencial (solo si tiene datos) ---
    asist = [
        ("Procedimientos", record.get("procedimientos", "")),
        ("Medicacion", record.get("medicacion_administrada", "")),
        ("Respuesta", record.get("respuesta", "")),
        ("Obs. legales", record.get("observaciones_legales", "")),
    ]
    if any(_v(v) for _, v in asist):
        _sec(pdf, "Parte asistencial")
        for lbl, val in asist:
            _row(pdf, lbl, val)

    # --- Firma ---
    if firma_bytes:
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                tmp.write(firma_bytes)
                tmp_path = tmp.name
            if pdf.get_y() > 240:
                pdf.add_page()
            _sec(pdf, "Firma profesional")
            y_img = pdf.get_y()
            pdf.image(tmp_path, x=14, y=y_img, w=50)
            pdf.set_xy(70, y_img + 4)
            pdf.set_font("Arial", "B", 8)
            pdf.cell(0, 5, safe_text(prof_nombre), ln=True)
            pdf.set_x(70)
            pdf.set_font("Arial", "", 8)
            pdf.cell(0, 5, safe_text(f"Mat: {prof_mat}"), ln=True)
            pdf.ln(20)
        except Exception as _exc:
            log_event("clinical_exports", f"fallo_firma_profesional:{type(_exc).__name__}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)

    # --- Pie de firma ---
    y_base = max(pdf.get_y() + 8, 250)
    if y_base > 270:
        pdf.add_page()
        y_base = 250
    pdf.line(14, y_base, 80, y_base)
    pdf.set_xy(14, y_base + 1)
    pdf.set_font("Arial", "B", 8)
    pdf.cell(66, 4, safe_text(prof_nombre), align="C", ln=True)
    pdf.set_x(14)
    pdf.set_font("Arial", "", 8)
    pdf.cell(66, 4, safe_text(f"Mat: {prof_mat}"), align="C", ln=True)

    pdf.line(110, y_base, 196, y_base)
    pdf.set_xy(110, y_base + 1)
    pdf.set_font("Arial", "B", 8)
    pdf.cell(86, 4, safe_text("Recepcion / conformidad"), align="C", ln=True)
    pdf.set_x(110)
    pdf.set_font("Arial", "", 8)
    familiar = _v(record.get("familiar_notificado", "")) or "_______________"
    pdf.cell(86, 4, safe_text(familiar), align="C", ln=True)

    return pdf_output_bytes(pdf)
