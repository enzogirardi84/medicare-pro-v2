"""Exportación a Excel y JSON del historial del paciente.

Extraído de core/clinical_exports.py.
"""
import io
import json

import pandas as pd

from core.utils import mapa_detalles_pacientes
from core._exports_helpers import collect_patient_sections


def build_patient_excel_bytes(session_state, paciente_sel):
    from datetime import datetime as _dt

    detalles = mapa_detalles_pacientes(session_state).get(paciente_sel, {})
    nom_pac = paciente_sel.split(" - ")[0] if " - " in paciente_sel else paciente_sel

    output = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        engine = "openpyxl"
    except Exception as e_openpyxl:
        try:
            import xlsxwriter  # noqa: F401
            engine = "xlsxwriter"
        except Exception as e_xlsx:
            # MEJORA: Registro explícito del fallo de las librerías
            from core.app_logging import log_event
            log_event("export_error", f"Fallo motores Excel. openpyxl: {e_openpyxl} | xlsxwriter: {e_xlsx}")
            return None

    sheets = {}
    _all_sections = collect_patient_sections(session_state, paciente_sel)
    _total_records = sum(len(v) for v in _all_sections.values())
    _non_empty = sum(1 for v in _all_sections.values() if v)
    from core.app_logging import log_event
    log_event("excel_export", f"Paciente={paciente_sel[:40]} secciones={len(_all_sections)} no_vacias={_non_empty} total_registros={_total_records}")

    # Fallback: si patient matching retorno vacio, leer directo de session state
    _DIRECT_KEYS = [
        ("checkin_db", "Auditoria de Presencia"),
        ("agenda_db", "Visitas y Agenda"),
        ("emergencias_db", "Emergencias y Ambulancia"),
        ("cuidados_enfermeria_db", "Enfermeria y Plan de Cuidados"),
        ("escalas_clinicas_db", "Escalas Clinicas"),
        ("auditoria_legal_db", "Auditoria Legal"),
        ("evoluciones_db", "Procedimientos y Evoluciones"),
        ("estudios_db", "Estudios Complementarios"),
        ("consumos_db", "Materiales Utilizados"),
        ("fotos_heridas_db", "Registro de Heridas"),
        ("vitales_db", "Signos Vitales"),
        ("pediatria_db", "Control Percentilo"),
        ("balance_db", "Balance Hidrico"),
        ("indicaciones_db", "Plan Terapeutico"),
        ("consentimientos_db", "Consentimientos"),
        ("facturacion_db", "Cobros y Facturacion"),
    ]
    for section_name, records in _all_sections.items():
        if not records:
            continue
        df = pd.DataFrame(records).drop(columns=["paciente", "empresa", "firma_b64"], errors="ignore")
        sheets[section_name[:31]] = df

    if not sheets:
        log_event("excel_export", f"SIN_DATOS paciente={paciente_sel[:40]} intentando fallback directo")
        for _key, _label in _DIRECT_KEYS:
            _raw = session_state.get(_key, [])
            if _raw:
                try:
                    _df = pd.DataFrame(_raw).drop(columns=["paciente", "empresa", "firma_b64"], errors="ignore")
                    sheets[_label[:31]] = _df
                except Exception:
                    continue
        if not sheets:
            sheets["Sin registros"] = pd.DataFrame([{"mensaje": "No hay registros clinicos para este paciente."}])

    with pd.ExcelWriter(output, engine=engine) as writer:
        resumen_rows = [{"Seccion": sec, "Registros": len(df)} for sec, df in sheets.items()]
        pd.DataFrame(resumen_rows).to_excel(writer, index=False, sheet_name="Resumen")
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        if engine == "openpyxl":
            wb = writer.book
            _navy = "162644"
            _green = "0D5A50"
            _hdr_font = Font(bold=True, color="FFFFFF", size=10)
            _hdr_fill_navy = PatternFill("solid", fgColor=_navy)
            _hdr_fill_green = PatternFill("solid", fgColor=_green)
            _thin = Side(style="thin", color="D1D5DB")
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

            for ws in wb.worksheets:
                fill = _hdr_fill_green if ws.title == "Resumen" else _hdr_fill_navy
                for cell in ws[1]:
                    cell.font = _hdr_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = _border
                ws.row_dimensions[1].height = 22
                for col in ws.columns:
                    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 50)
                alt_fill = PatternFill("solid", fgColor="F3F4F6")
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    row_fill = alt_fill if i % 2 == 0 else None
                    for cell in row:
                        cell.border = _border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        if row_fill:
                            cell.fill = row_fill
                ws.freeze_panes = "A2"

            ws_res = wb["Resumen"]
            ws_res.insert_rows(1, amount=4)
            ws_res["A1"] = detalles.get("empresa", "")
            ws_res["A1"].font = Font(bold=True, size=14, color=_navy)
            ws_res["A2"] = "HISTORIA CLINICA DIGITAL - EXPORTACION COMPLETA"
            ws_res["A2"].font = Font(bold=True, size=11, color=_navy)
            ws_res["A3"] = f"Paciente: {nom_pac}   |   DNI: {detalles.get('dni', 'S/D')}   |   Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
            ws_res["A3"].font = Font(size=9, color="64748B")
            ws_res["A4"] = ""

    output.seek(0)
    return output.getvalue()


def build_patient_json_bytes(session_state, paciente_sel):
    payload = {
        "paciente": paciente_sel,
        "detalles": mapa_detalles_pacientes(session_state).get(paciente_sel, {}),
        "secciones": collect_patient_sections(session_state, paciente_sel),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")
