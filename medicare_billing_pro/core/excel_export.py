"""Exportacion a Excel para Medicare Billing Pro."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Sequence

from core.utils import fmt_fecha

XLSX_DISPONIBLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    XLSX_DISPONIBLE = True
except ImportError:
    pass

BRAND = "1F6FEB"
INK = "0F172A"
MUTED = "64748B"
HEADER = "0F172A"
HEADER_2 = "E2E8F0"
TOTAL_FILL = "DCFCE7"
ALT_FILL = "F8FAFC"
MONEY_FORMAT = '"$"#,##0.00'
DATE_FORMAT = "DD/MM/YYYY"


def _workbook() -> "Workbook":
    wb = Workbook()
    wb.properties.creator = "Medicare Billing Pro"
    wb.properties.company = "Medicare Pro Suite"
    wb.properties.title = "Medicare Billing Pro"
    wb.properties.created = datetime.now()
    return wb


def _safe_sheet_name(name: str) -> str:
    return str(name or "Hoja")[:31].replace("/", "-").replace("\\", "-").replace("*", "-")


def _write_title(ws, title: str, empresa: str, subtitle: str = "") -> int:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=INK)
    ws["A2"] = f"{empresa} | Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Aptos", size=9, color=MUTED)
    if subtitle:
        ws["A3"] = subtitle
        ws["A3"].font = Font(name="Aptos", size=9, color=MUTED)
        return 5
    return 4


def _style_header(ws, headers: Sequence[str], row: int) -> None:
    fill = PatternFill(start_color=HEADER, end_color=HEADER, fill_type="solid")
    font = Font(name="Aptos", bold=True, color="FFFFFF", size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions


def _thin_border():
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _append_rows(ws, headers: Sequence[str], rows: Iterable[Sequence[Any]], start_row: int) -> int:
    _style_header(ws, headers, start_row)
    row_idx = start_row + 1
    border = _thin_border()
    for row in rows:
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=ALT_FILL, end_color=ALT_FILL, fill_type="solid")
        row_idx += 1
    return row_idx - 1


def _add_table(ws, name: str, header_row: int, last_row: int, last_col: int) -> None:
    if last_row <= header_row:
        return
    ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _format_money_cols(ws, columns: Sequence[int], first_row: int, last_row: int) -> None:
    for col in columns:
        for row in range(first_row, last_row + 1):
            ws.cell(row=row, column=col).number_format = MONEY_FORMAT
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="top")


def _format_date_cols(ws, columns: Sequence[int], first_row: int, last_row: int) -> None:
    for col in columns:
        for row in range(first_row, last_row + 1):
            ws.cell(row=row, column=col).number_format = DATE_FORMAT


def _auto_width(ws, min_width: int = 11, max_width: int = 44) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _finish(wb: "Workbook") -> bytes:
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _total_row(ws, row: int, label_col: int, value_col: int, label: str, value: float) -> None:
    ws.cell(row=row, column=label_col, value=label)
    ws.cell(row=row, column=value_col, value=value)
    for col in range(label_col, value_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Aptos", bold=True, color=INK)
        cell.fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")
        cell.border = _thin_border()
    ws.cell(row=row, column=value_col).number_format = MONEY_FORMAT


def exportar_clientes_excel(clientes: List[Dict], empresa: str) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Clientes"
    header_row = _write_title(ws, "Clientes fiscales", empresa, f"{len(clientes)} registros")
    headers = ["Nombre / Razon Social", "DNI / CUIT", "Email", "Telefono", "Direccion", "Condicion Fiscal", "Notas"]
    rows = [
        [
            c.get("nombre", ""),
            c.get("dni", ""),
            c.get("email", ""),
            c.get("telefono", ""),
            c.get("direccion", ""),
            c.get("condicion_fiscal", ""),
            c.get("notas", ""),
        ]
        for c in clientes
    ]
    last = _append_rows(ws, headers, rows, header_row)
    _add_table(ws, "tbl_clientes", header_row, last, len(headers))
    _auto_width(ws)
    return _finish(wb)


def exportar_cobros_excel(cobros: List[Dict], empresa: str) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Cobros"
    total = sum(float(c.get("monto", 0) or 0) for c in cobros)
    header_row = _write_title(ws, "Historial de cobros", empresa, f"{len(cobros)} registros | Total: ${total:,.2f}")
    headers = ["Fecha", "Cliente", "Concepto", "Metodo de Pago", "Monto", "Estado", "Notas"]
    rows = [
        [
            fmt_fecha(c.get("fecha", "")),
            c.get("cliente_nombre", ""),
            c.get("concepto", ""),
            c.get("metodo_pago", ""),
            float(c.get("monto", 0) or 0),
            c.get("estado", ""),
            c.get("notas", ""),
        ]
        for c in cobros
    ]
    last = _append_rows(ws, headers, rows, header_row)
    if last >= header_row + 1:
        _format_money_cols(ws, [5], header_row + 1, last)
        _total_row(ws, last + 2, 4, 5, "Total cobrado", total)
    _add_table(ws, "tbl_cobros", header_row, last, len(headers))
    _auto_width(ws)
    return _finish(wb)


def exportar_estado_cuenta_excel(
    cliente: Dict[str, Any],
    empresa: str,
    movimientos: List[Dict[str, Any]],
    total_debe: float,
    total_haber: float,
    saldo: float,
) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Estado de cuenta"
    subtitle = (
        f"Cliente: {cliente.get('nombre', '')} | "
        f"Debe: ${total_debe:,.2f} | Haber: ${total_haber:,.2f} | Saldo: ${saldo:,.2f}"
    )
    header_row = _write_title(ws, "Estado de cuenta", empresa, subtitle)
    headers = ["Fecha", "Tipo", "Numero", "Detalle", "Debe", "Haber", "Saldo"]
    rows = [
        [
            fmt_fecha(m.get("fecha", "")),
            m.get("tipo", ""),
            m.get("numero", ""),
            m.get("detalle", ""),
            float(m.get("debe", 0) or 0),
            float(m.get("haber", 0) or 0),
            float(m.get("saldo", 0) or 0),
        ]
        for m in movimientos
    ]
    last = _append_rows(ws, headers, rows, header_row)
    if last >= header_row + 1:
        _format_money_cols(ws, [5, 6, 7], header_row + 1, last)
        _total_row(ws, last + 2, 4, 5, "Debe", total_debe)
        _total_row(ws, last + 3, 5, 6, "Haber", total_haber)
        _total_row(ws, last + 4, 6, 7, "Saldo", saldo)
    _add_table(ws, "tbl_estado_cuenta", header_row, last, len(headers))
    _auto_width(ws)
    return _finish(wb)


def exportar_presupuestos_excel(presupuestos: List[Dict], empresa: str) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Presupuestos"
    total = sum(float(p.get("total", 0) or 0) for p in presupuestos)
    header_row = _write_title(ws, "Presupuestos", empresa, f"{len(presupuestos)} documentos | Total: ${total:,.2f}")
    headers = ["Numero", "Fecha", "Cliente", "Total", "Estado", "Valido hasta", "Notas"]
    rows = [
        [
            p.get("numero", ""),
            fmt_fecha(p.get("fecha", "")),
            p.get("cliente_nombre", ""),
            float(p.get("total", 0) or 0),
            p.get("estado", ""),
            fmt_fecha(p.get("valido_hasta", "")),
            p.get("notas", ""),
        ]
        for p in presupuestos
    ]
    last = _append_rows(ws, headers, rows, header_row)
    if last >= header_row + 1:
        _format_money_cols(ws, [4], header_row + 1, last)
        _total_row(ws, last + 2, 3, 4, "Total presupuestado", total)
    _add_table(ws, "tbl_presupuestos", header_row, last, len(headers))
    _auto_width(ws)
    return _finish(wb)


def exportar_prefacturas_excel(prefacturas: List[Dict], empresa: str) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Pre-facturas"
    total = sum(float(p.get("total", 0) or 0) for p in prefacturas)
    header_row = _write_title(ws, "Pre-facturas", empresa, f"{len(prefacturas)} documentos | Total: ${total:,.2f}")
    headers = ["Numero", "Fecha", "Cliente", "DNI / CUIT", "Total", "Cobrado", "Saldo", "Estado", "Notas"]
    rows = [
        [
            p.get("numero", ""),
            fmt_fecha(p.get("fecha", "")),
            p.get("cliente_nombre", ""),
            p.get("cliente_dni", ""),
            float(p.get("total", 0) or 0),
            float(p.get("cobrado", 0) or 0),
            float(p.get("saldo", p.get("total", 0)) or 0),
            p.get("estado_calculado", p.get("estado", "")),
            p.get("notas", ""),
        ]
        for p in prefacturas
    ]
    last = _append_rows(ws, headers, rows, header_row)
    if last >= header_row + 1:
        _format_money_cols(ws, [5, 6, 7], header_row + 1, last)
        _total_row(ws, last + 2, 4, 5, "Total pre-facturado", total)
        _total_row(ws, last + 3, 6, 7, "Saldo pendiente", sum(float(p.get("saldo", p.get("total", 0)) or 0) for p in prefacturas))
    _add_table(ws, "tbl_prefacturas", header_row, last, len(headers))
    _auto_width(ws)
    return _finish(wb)


def _summary_sheet(ws, empresa: str, mes: str, presupuestos: List[Dict], prefacturas: List[Dict], cobros: List[Dict]) -> None:
    total_presupuestado = sum(float(p.get("total", 0) or 0) for p in presupuestos)
    total_facturado = sum(float(p.get("total", 0) or 0) for p in prefacturas)
    total_cobrado = sum(float(c.get("monto", 0) or 0) for c in cobros)
    pendiente = total_facturado - total_cobrado
    row = _write_title(ws, f"Reporte contable - {mes}", empresa)
    headers = ["Indicador", "Monto", "Cantidad"]
    rows = [
        ["Presupuestado", total_presupuestado, len(presupuestos)],
        ["Pre-facturado", total_facturado, len(prefacturas)],
        ["Cobrado", total_cobrado, len(cobros)],
        ["Pendiente", pendiente, ""],
    ]
    last = _append_rows(ws, headers, rows, row)
    _format_money_cols(ws, [2], row + 1, last)
    for r in range(row + 1, last + 1):
        ws.cell(r, 1).font = Font(name="Aptos", bold=True, color=INK)
    _auto_width(ws)


def exportar_facturas_arca_excel(facturas: List[Dict], empresa: str) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Facturas ARCA"
    total = sum(float(f.get("total", 0) or 0) for f in facturas)
    header_row = _write_title(ws, "Facturas ARCA", empresa, f"{len(facturas)} comprobantes | Total: ${total:,.2f}")
    headers = [
        "Numero",
        "PV",
        "Tipo",
        "Fecha",
        "Cliente",
        "CUIT/DNI",
        "Condicion IVA",
        "Neto",
        "IVA",
        "Total",
        "Estado",
        "CAE",
        "Vto CAE",
        "Notas",
    ]
    rows = [
        [
            f.get("numero", ""),
            str(f.get("punto_venta", 1)).zfill(4),
            f.get("tipo_comprobante", "C"),
            fmt_fecha(f.get("fecha", "")),
            f.get("cliente_nombre", ""),
            f.get("cliente_dni", ""),
            f.get("condicion_iva_receptor", ""),
            float(f.get("neto", 0) or 0),
            float(f.get("iva", 0) or 0),
            float(f.get("total", 0) or 0),
            f.get("estado", ""),
            f.get("cae", ""),
            fmt_fecha(f.get("cae_vencimiento", "")) if f.get("cae_vencimiento") else "",
            f.get("notas", ""),
        ]
        for f in facturas
    ]
    last = _append_rows(ws, headers, rows, header_row)
    if last >= header_row + 1:
        _format_money_cols(ws, [8, 9, 10], header_row + 1, last)
        _total_row(ws, last + 2, 7, 8, "Neto", sum(float(f.get("neto", 0) or 0) for f in facturas))
        _total_row(ws, last + 3, 8, 9, "IVA", sum(float(f.get("iva", 0) or 0) for f in facturas))
        _total_row(ws, last + 4, 9, 10, "Total", total)
    _add_table(ws, "tbl_facturas_arca", header_row, last, len(headers))

    ws2 = wb.create_sheet("Detalle conceptos")
    detail_header = _write_title(ws2, "Detalle conceptos ARCA", empresa)
    detail_headers = ["Factura", "Fecha", "Cliente", "Concepto", "Cantidad", "Precio unitario", "Subtotal"]
    detail_rows = []
    for f in facturas:
        for item in f.get("items", []) or []:
            cantidad = float(item.get("cantidad", 1) or 1)
            precio = float(item.get("precio_unitario", 0) or 0)
            detail_rows.append([f.get("numero", ""), fmt_fecha(f.get("fecha", "")), f.get("cliente_nombre", ""), item.get("concepto", ""), cantidad, precio, cantidad * precio])
    detail_last = _append_rows(ws2, detail_headers, detail_rows, detail_header)
    if detail_last >= detail_header + 1:
        _format_money_cols(ws2, [6, 7], detail_header + 1, detail_last)
    _add_table(ws2, "tbl_facturas_arca_items", detail_header, detail_last, len(detail_headers))
    _auto_width(ws)
    _auto_width(ws2)
    return _finish(wb)


def exportar_reporte_contable_excel(
    empresa: str,
    mes: str,
    presupuestos: List[Dict],
    prefacturas: List[Dict],
    cobros: List[Dict],
) -> bytes:
    if not XLSX_DISPONIBLE:
        return b""
    wb = _workbook()
    ws = wb.active
    ws.title = "Resumen"
    _summary_sheet(ws, empresa, mes, presupuestos, prefacturas, cobros)

    ws2 = wb.create_sheet("Cobros")
    header = _write_title(ws2, "Cobros del mes", empresa, mes)
    rows = [
        [fmt_fecha(c.get("fecha", "")), c.get("cliente_nombre", ""), c.get("concepto", ""), c.get("metodo_pago", ""), float(c.get("monto", 0) or 0), c.get("estado", "")]
        for c in cobros
    ]
    last = _append_rows(ws2, ["Fecha", "Cliente", "Concepto", "Metodo", "Monto", "Estado"], rows, header)
    if last >= header + 1:
        _format_money_cols(ws2, [5], header + 1, last)
    _add_table(ws2, "tbl_rep_cobros", header, last, 6)
    _auto_width(ws2)

    ws3 = wb.create_sheet("Pre-facturas")
    header = _write_title(ws3, "Pre-facturas del mes", empresa, mes)
    rows = [
        [
            p.get("numero", ""),
            fmt_fecha(p.get("fecha", "")),
            p.get("cliente_nombre", ""),
            float(p.get("total", 0) or 0),
            float(p.get("cobrado", 0) or 0),
            float(p.get("saldo", p.get("total", 0)) or 0),
            p.get("estado_calculado", p.get("estado", "")),
        ]
        for p in prefacturas
    ]
    last = _append_rows(ws3, ["Numero", "Fecha", "Cliente", "Total", "Cobrado", "Saldo", "Estado"], rows, header)
    if last >= header + 1:
        _format_money_cols(ws3, [4, 5, 6], header + 1, last)
    _add_table(ws3, "tbl_rep_prefacturas", header, last, 7)
    _auto_width(ws3)

    ws4 = wb.create_sheet("Presupuestos")
    header = _write_title(ws4, "Presupuestos del mes", empresa, mes)
    rows = [
        [p.get("numero", ""), fmt_fecha(p.get("fecha", "")), p.get("cliente_nombre", ""), float(p.get("total", 0) or 0), p.get("estado", "")]
        for p in presupuestos
    ]
    last = _append_rows(ws4, ["Numero", "Fecha", "Cliente", "Total", "Estado"], rows, header)
    if last >= header + 1:
        _format_money_cols(ws4, [4], header + 1, last)
    _add_table(ws4, "tbl_rep_presupuestos", header, last, 5)
    _auto_width(ws4)

    return _finish(wb)
